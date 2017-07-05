import csv
import numpy as np
from openpyxl import load_workbook
from scipy.stats import chi2, norm

def load_return_data():
	# Totally 46 missing data points
	sheets = ['Size', 'Gross Profitability', 'Value', 'ValProf', 'Accruals', \
		'Net Issuance (rebal.-A)', 'Asset Growth', 'Investment', 'Piotroski\'s F-score', \
		'ValMomProf', 'ValMom', 'Idiosyncratic Volatility', 'Momentum', 'Long Run Reversals', 'Beta Arbitrage']
	wb = load_workbook('data/Simple_Strategies_returns.xlsx')
	returns = list()
	for sheet in sheets:
		ws = wb[sheet]
		r = []
		for row in ws.iter_rows():
			r.append([cell.value if cell.value is not None else 0 for cell in row[1:]])
		returns.append(np.array(r[1:]))

	# Totally 96 missing data points
	r = []
	with open('data/49_Industry_Portfolios.CSV', 'rb') as csvfile:
		reader = csv.reader(csvfile)
		cur_row = 0
		for row in reader:
			cur_row += 1
			if cur_row < 457:
				continue
			if cur_row > 1062:
				break
			r.append([float(element) for element in row[1:]])
	returns.append(np.array(r))
	missing_data = (returns[-1] == -99.99)
	returns[-1][missing_data] = 0
	return np.concatenate(returns, 1).T

def load_factor_data():
	# 1 missing data point
	factor_name = ['Size', 'Gross Profitability', 'Value', 'ValProf', 'Accruals', \
		'Net Issuance (rebal.:A)', 'Asset Growth', 'Investment', 'Piotroski\'s F-score', \
		'ValMomProf', 'ValMom', 'Idiosyncratic Volatility', 'Momentum', 'Long Run Reversals', 'Beta Arbitrage']
	wb = load_workbook('data/Simple_Strategies_returns.xlsx')
	ws = wb['Long-Short Gross Returns']
	first_row = True
	columns = []
	factors = []
	f = []
	for row in ws.iter_rows():
		if first_row:
			first_row = False
			for i in range(len(row)):
				if row[i].value in factor_name:
					columns.append(i)
		else:
			f.append([row[col].value if row[col].value is not None else 0 for col in columns])
	factors.append(np.array(f))

	with open('data/F-F_Research_Data_5_Factors_2x3.CSV', 'rb') as csvfile:
		reader = csv.reader(csvfile)
		cur_row = 0
		f = []
		for row in reader:
			cur_row += 1
			if cur_row < 5:
				continue
			if cur_row > 610:
				break
			f.append([float(e) for e in row[1:]])
		factors.append(np.array(f))
	return np.concatenate(factors, 1).T

def test_statistic(F, Y, T, N, Q):
	"""Test statistic for time-series regression

	Args: 
		F: A Q * T matrix representing factors
		Y: An N * T matrix representing asset returns
		N: Number of assets
		T: Number of (time) observations
		Q: Number of factors

	Returns: 
		Lambda: Test statistic
		SD: Standard deviation
		CV: Confidence interval
		p_value: P-value
	"""

	y = 1.0 * N / T
	X = np.concatenate((np.ones((1,T)),F), axis = 0)
	B_hat = np.dot(np.dot(Y, np.transpose(X)), np.linalg.inv(np.dot(X, np.transpose(X))))
	alpha_hat = B_hat[:,0]
	Sigma_hat = (np.dot(Y, np.transpose(Y)) - np.dot(np.dot(B_hat, X), np.transpose(Y))) / T
	k = T - np.dot(np.sum(F, axis = 1), np.dot(np.linalg.inv(np.dot(F, np.transpose(F))), np.sum(F, axis = 1)))

	Lambda = 1.0 * (T - N - Q) / T / N * k * np.dot(alpha_hat, np.dot(np.linalg.inv(Sigma_hat), alpha_hat)) # test statistic
	SD = np.sqrt(2.0 / (1 - y) / N) # standard error
	CV = (Lambda - 1) / SD # critical value
	p_value = 1 - norm.cdf(CV) # p-value
	return (Lambda, SD, CV, p_value)

def test_statistic_cross(F, Y, T, N, Q):
	"""Test statistic for cross-sectional regression

	Args: 
		F: A Q * T matrix representing factors
		Y: An N * T matrix representing asset returns
		N: Number of assets
		T: Number of (time) observations
		Q: Number of factors

	Returns: 
		Lambda: Test statistic
		SD: Standard deviation
		CV: Confidence interval
		p_value: P-value
	"""

	y = 1.0 * N / T
	X = np.concatenate((np.ones((1,T)),F), axis = 0)
	B_hat = np.dot(np.dot(Y, np.transpose(X)), np.linalg.inv(np.dot(X, np.transpose(X))))
	beta = B_hat[:,1:]
	P1 = np.identity(T) - np.dot(np.dot(np.transpose(F), np.linalg.pinv(np.dot(F, np.transpose(F)))), F)
	P2 = np.identity(N) - np.dot(beta, np.dot(np.linalg.pinv(np.dot(np.transpose(beta), beta)), np.transpose(beta)))
	Sigma = np.dot(Y, np.dot(P1, np.transpose(Y))) / T
	alpha = np.dot(P2, np.sum(Y, axis = 1) / T)
	cov = np.dot(np.dot(P2, Sigma),P2) / T
	F_bar = np.sum(F, axis = 1) / T
	k = (1 - np.dot(F_bar, np.dot(np.linalg.pinv(np.dot(F, np.transpose(F)) / T), F_bar))) * T
	Lambda = np.dot(alpha, np.dot(np.linalg.pinv(cov), alpha)) * (T - Q) * k / (N - Q) / T / T
	SD = np.sqrt(2.0 * (1 - y) / (N - Q)) # standard error
	CV = (Lambda - 1) / SD # critical value
	p_value = 1 - norm.cdf(CV) # p-value
	return (Lambda, SD, CV, p_value)

def result(F, Y, T, N, Q):
	Lambda, SD, CV, p_value = test_statistic(F, Y, T, N, Q) # Our test statistic
	Lambda_1 = 1.0 * T / (T - N - Q) * N * Lambda # test statistic 1
	p_value_1 = 1 - chi2.cdf(Lambda_1, N)
	Lambda_2 = N * Lambda # test statistic 2
	p_value_2 = 1 - chi2.cdf(Lambda_2, N)
	return (Lambda, p_value, Lambda_1, p_value_1, Lambda_2, p_value_2)

def result_cross(F, Y, T, N, Q):
	Lambda, SD, CV, p_value = test_statistic_cross(F, Y, T, N, Q)
	Lambda_1 = (N - Q) * Lambda
	p_value_1 = 1 - chi2.cdf(Lambda_1, N - Q)
	return (Lambda, p_value, Lambda_1, p_value_1)

def display(result, message, N, T, Q, f = None):
	if f is None:
		print '%' * 80
		print '    %d Portfolios (Novy Marx + Kenneth Industry): N = %d, T = %d' % (N, N, T)
		print '    Use %s Data From %s To %s. ' % (message[0], message[1], message[2])
		print '    Test On %d Factors. \n' % (Q)
		print '{:>20}  {:>20}  {:>20}'.format('Test Statistic', 'Estimation', 'p-value')
		print '{:>20}  {:>20}  {:>20}'.format('Lambda', str(result[0]), str(result[1]))
		print '{:>20}  {:>20}  {:>20}'.format('Lambda_1', str(result[2]), str(result[3]))
		print '{:>20}  {:>20}  {:>20}'.format('Lambda_2', str(result[4]), str(result[5]))
		print '\n' + '%' * 80 + '\n'
	else:
		f.write('%' * 80 + '\n')
		f.write('    %d Portfolios (Novy Marx + Kenneth Industry): N = %d, T = %d\n' % (N, N, T))
		f.write('    Use %s Data From %s To %s. \n' % (message[0], message[1], message[2]))
		f.write('    Test On %d Factors. \n\n' % (Q))
		f.write('{:>20}  {:>20}  {:>20}\n'.format('Test Statistic', 'Estimation', 'p-value'))
		f.write('{:>20}  {:>20}  {:>20}\n'.format('Lambda', str(result[0]), str(result[1])))
		f.write('{:>20}  {:>20}  {:>20}\n'.format('Lambda_1', str(result[2]), str(result[3])))
		f.write('{:>20}  {:>20}  {:>20}\n'.format('Lambda_2', str(result[4]), str(result[5])))
		f.write('\n' + '%' * 80 + '\n\n')

def display_cross(result, message, N, T, Q, f = None):
	if f is None:
		print '%' * 80
		print '    %d Portfolios (Novy Marx + Kenneth Industry): N = %d, T = %d' % (N, N, T)
		print '    Use %s Data From %s To %s. ' % (message[0], message[1], message[2])
		print '    Test On %d Factors. \n' % (Q)
		print '{:>20}  {:>20}  {:>20}'.format('Test Statistic', 'Estimation', 'p-value')
		print '{:>20}  {:>20}  {:>20}'.format('Lambda', str(result[0]), str(result[1]))
		print '{:>20}  {:>20}  {:>20}'.format('Lambda_1', str(result[2]), str(result[3]))
		print '\n' + '%' * 80 + '\n'
	else:
		f.write('%' * 80 + '\n')
		f.write('    %d Portfolios (Novy Marx + Kenneth Industry): N = %d, T = %d\n' % (N, N, T))
		f.write('    Use %s Data From %s To %s. \n' % (message[0], message[1], message[2]))
		f.write('    Test On %d Factors. \n\n' % (Q))
		f.write('{:>20}  {:>20}  {:>20}\n'.format('Test Statistic', 'Estimation', 'p-value'))
		f.write('{:>20}  {:>20}  {:>20}\n'.format('Lambda', str(result[0]), str(result[1])))
		f.write('{:>20}  {:>20}  {:>20}\n'.format('Lambda_1', str(result[2]), str(result[3])))
		f.write('\n' + '%' * 80 + '\n\n')